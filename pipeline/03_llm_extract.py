# Author: Samuel Ahuno
# Date: 2026-04-20
# Purpose: Extract LINE-1 somatic insertion loci from papers using the Claude API

import argparse
import base64
import csv
import io
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MODEL_PDF = "claude-opus-4-5"
MODEL_ABSTRACT = "claude-haiku-4-5-20251001"

# Token cost estimates (per million tokens, USD) for logging only
COST_PER_MTOK = {
    MODEL_PDF: {"input": 15.0, "output": 75.0},
    MODEL_ABSTRACT: {"input": 0.80, "output": 4.0},
}

MAX_TOKENS_RESPONSE = 8192

REPAIR_PROMPT = (
    "Your previous response could not be parsed as JSON. "
    "Return ONLY a valid JSON object matching the schema: "
    '{"loci": [...], "reason": "..."}. '
    "No markdown, no prose. Only raw JSON."
)


# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

def setup_logging(log_dir: Path, script_name: str) -> logging.Logger:
    """
    Configure dual-handler logging (file + stdout).

    Parameters
    ----------
    log_dir : Path
        Directory where the log file will be written.
    script_name : str
        Used as the logger name and log file prefix.

    Returns
    -------
    logging.Logger
        Configured logger instance.
    """
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"{script_name}_{timestamp}.log"

    logger = logging.getLogger(script_name)
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s")

    fh = logging.FileHandler(log_file)
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


# ---------------------------------------------------------------------------
# Claude API helpers
# ---------------------------------------------------------------------------

def load_system_prompt(prompt_path: Path, logger: logging.Logger) -> str:
    """
    Load the extraction system prompt from disk.

    Parameters
    ----------
    prompt_path : Path
        Path to the plain-text prompt file.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    str
        Prompt text.

    Raises
    ------
    SystemExit
        If the file does not exist.
    """
    if not prompt_path.exists():
        logger.error("Extraction prompt not found: %s", prompt_path.resolve())
        sys.exit(1)
    text = prompt_path.read_text(encoding="utf-8").strip()
    logger.info("Loaded extraction prompt: %s (%d chars)", prompt_path.resolve(), len(text))
    return text


def _build_messages_pdf(pdf_bytes: bytes, system_prompt: str) -> list[dict]:
    """
    Build the Anthropic messages payload for a PDF document.

    Parameters
    ----------
    pdf_bytes : bytes
        Raw PDF file content.
    system_prompt : str
        Extraction instructions (unused here — passed as system param to API).

    Returns
    -------
    list[dict]
        Messages list for the Anthropic API.
    """
    b64 = base64.standard_b64encode(pdf_bytes).decode("ascii")
    return [
        {
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": b64,
                    },
                },
                {
                    "type": "text",
                    "text": "Extract all somatic LINE-1 insertion loci from this paper. Return only valid JSON.",
                },
            ],
        }
    ]


def _build_messages_abstract(abstract: str) -> list[dict]:
    """
    Build the Anthropic messages payload for an abstract-only paper.

    Parameters
    ----------
    abstract : str
        Paper abstract text.

    Returns
    -------
    list[dict]
        Messages list for the Anthropic API.
    """
    return [
        {
            "role": "user",
            "content": (
                f"ABSTRACT:\n{abstract}\n\n"
                "Extract all somatic LINE-1 insertion loci reported in this abstract. "
                "Return only valid JSON."
            ),
        }
    ]


def _build_system_block(system_prompt: str) -> list[dict]:
    """
    Wrap the system prompt with cache_control for prompt caching.

    Parameters
    ----------
    system_prompt : str
        The extraction instructions text.

    Returns
    -------
    list[dict]
        System content list with cache_control applied to the prompt block.

    Example
    -------
    >>> blocks = _build_system_block("Extract loci...")
    >>> blocks[0]["cache_control"]
    {'type': 'ephemeral'}
    """
    return [
        {
            "type": "text",
            "text": system_prompt,
            "cache_control": {"type": "ephemeral"},
        }
    ]


def call_claude(
    client: anthropic.Anthropic,
    model: str,
    system_blocks: list[dict],
    messages: list[dict],
    logger: logging.Logger,
) -> anthropic.types.Message:
    """
    Call the Claude API with retry on transient errors.

    Parameters
    ----------
    client : anthropic.Anthropic
        Authenticated Anthropic client.
    model : str
        Model identifier string.
    system_blocks : list[dict]
        System prompt content blocks.
    messages : list[dict]
        User/assistant messages.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    anthropic.types.Message
        API response object.

    Raises
    ------
    anthropic.APIError
        After exhausting retries.
    """
    max_retries = 4
    backoff = 5.0
    for attempt in range(1, max_retries + 1):
        try:
            return client.messages.create(
                model=model,
                max_tokens=MAX_TOKENS_RESPONSE,
                system=system_blocks,
                messages=messages,
            )
        except anthropic.RateLimitError as exc:
            wait = backoff * (2 ** (attempt - 1))
            logger.warning(
                "Rate limit hit (attempt %d/%d). Sleeping %.0fs: %s",
                attempt, max_retries, wait, exc,
            )
            time.sleep(wait)
        except anthropic.APIStatusError as exc:
            if exc.status_code >= 500 and attempt < max_retries:
                wait = backoff * (2 ** (attempt - 1))
                logger.warning(
                    "Server error %d (attempt %d/%d). Sleeping %.0fs.",
                    exc.status_code, attempt, max_retries, wait,
                )
                time.sleep(wait)
            else:
                raise
    raise anthropic.APIError("Exhausted retries for Claude API call")


def extract_json(response_text: str) -> dict | None:
    """
    Attempt to parse a JSON object from the model response text.

    Strips markdown fences if present before parsing.

    Parameters
    ----------
    response_text : str
        Raw text content from the Claude response.

    Returns
    -------
    dict or None
        Parsed dict, or None if parsing fails.
    """
    text = response_text.strip()
    # Strip markdown code fences
    if text.startswith("```"):
        lines = text.splitlines()
        text = "\n".join(
            line for line in lines if not line.startswith("```")
        ).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def estimate_cost(usage: anthropic.types.Usage, model: str) -> float:
    """
    Estimate USD cost from an API usage object.

    Parameters
    ----------
    usage : anthropic.types.Usage
        Token usage from the API response.
    model : str
        Model identifier to look up rates.

    Returns
    -------
    float
        Estimated cost in USD.

    Example
    -------
    >>> estimate_cost(usage, "claude-haiku-4-5-20251001")
    0.0012
    """
    rates = COST_PER_MTOK.get(model, {"input": 0.0, "output": 0.0})
    input_cost = (usage.input_tokens / 1_000_000) * rates["input"]
    output_cost = (usage.output_tokens / 1_000_000) * rates["output"]
    return input_cost + output_cost


# ---------------------------------------------------------------------------
# Supplement parsing
# ---------------------------------------------------------------------------

# Max rows to include per sheet/file to keep token count manageable
MAX_ROWS_PER_SHEET = 2000


def _excel_to_text(path: Path, logger: logging.Logger) -> str:
    """
    Convert an Excel file to a plain-text CSV-like representation.

    Reads up to MAX_ROWS_PER_SHEET rows per sheet, skipping empty sheets.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_out: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_ROWS_PER_SHEET:
                    rows_out.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)")
                    break
                # Skip entirely empty rows
                if all(c is None for c in row):
                    continue
                rows_out.append("\t".join("" if c is None else str(c) for c in row))
            if rows_out:
                parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_out))
        wb.close()
        return "\n\n".join(parts)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to parse Excel %s: %s", path.name, exc)
        return ""


def _tabular_to_text(path: Path, logger: logging.Logger) -> str:
    """Convert a CSV or TSV file to plain text, capped at MAX_ROWS_PER_SHEET rows."""
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
        dialect = "excel-tab" if path.suffix.lower() == ".tsv" else "excel"
        reader = csv.reader(io.StringIO(raw), dialect=dialect)
        lines: list[str] = []
        for i, row in enumerate(reader):
            if i >= MAX_ROWS_PER_SHEET:
                lines.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)")
                break
            lines.append("\t".join(row))
        return "\n".join(lines)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to parse tabular file %s: %s", path.name, exc)
        return ""


def supplements_to_text(paths: list[Path], logger: logging.Logger) -> str:
    """
    Convert a list of supplement files to a single combined text block.

    Parameters
    ----------
    paths : list[Path]
        Supplement file paths (xlsx, xls, csv, tsv, txt).
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    str
        Combined text suitable for inclusion in a Claude message.
        Empty string if no usable content found.
    """
    sections: list[str] = []
    for p in paths:
        ext = p.suffix.lower()
        if ext in {".xlsx", ".xls"}:
            text = _excel_to_text(p, logger)
        elif ext in {".csv", ".tsv"}:
            text = _tabular_to_text(p, logger)
        elif ext == ".txt":
            try:
                text = p.read_text(encoding="utf-8", errors="replace")[:50_000]
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed to read txt %s: %s", p.name, exc)
                text = ""
        else:
            continue
        if text.strip():
            sections.append(f"--- Supplementary file: {p.name} ---\n{text.strip()}")
    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# Per-paper extraction
# ---------------------------------------------------------------------------

def process_paper(
    paper: dict,
    client: anthropic.Anthropic,
    system_prompt: str,
    extractions_dir: Path,
    logger: logging.Logger,
) -> tuple[list[dict], str, float]:
    """
    Run extraction for a single paper, returning loci, status, and cost.

    Parameters
    ----------
    paper : dict
        Paper metadata dict with keys pmid, abstract, pdf_path, pdf_available,
        supplement_paths.
    client : anthropic.Anthropic
        Authenticated Anthropic client.
    system_prompt : str
        Extraction instructions.
    extractions_dir : Path
        Directory to write per-paper raw extraction JSON.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    tuple[list[dict], str, float]
        (loci_list, extraction_status, cost_usd)
    """
    pmid = paper.get("pmid", "unknown")
    pdf_path_str = paper.get("pdf_path")
    pdf_available = paper.get("pdf_available", False)
    abstract = paper.get("abstract", "")
    supplement_path_strs: list[str] = paper.get("supplement_paths", [])

    system_blocks = _build_system_block(system_prompt)
    total_cost = 0.0

    # --- PDF ---
    pdf_bytes: bytes | None = None
    if pdf_available and pdf_path_str:
        pdf_path = Path(pdf_path_str)
        if pdf_path.exists():
            try:
                pdf_bytes = pdf_path.read_bytes()
                logger.debug("PMID %s: loaded PDF (%s)", pmid, pdf_path.name)
            except OSError as exc:
                logger.warning("PMID %s: cannot read PDF (%s), falling back", pmid, exc)
        else:
            logger.warning("PMID %s: pdf_path recorded but file missing", pmid)

    # --- Supplements ---
    supp_text = ""
    if supplement_path_strs:
        supp_paths = [Path(p) for p in supplement_path_strs if Path(p).exists()]
        if supp_paths:
            supp_text = supplements_to_text(supp_paths, logger)
            if supp_text:
                logger.info(
                    "PMID %s: parsed %d supplement file(s) → %d chars",
                    pmid, len(supp_paths), len(supp_text),
                )

    # --- Build message and choose model ---
    # Priority: PDF+supplements > supplements only > abstract only
    # Use opus whenever we have real document content; haiku for abstract-only
    if pdf_bytes is not None:
        model = MODEL_PDF
        content: list[dict] = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": base64.standard_b64encode(pdf_bytes).decode("ascii"),
                },
            },
        ]
        if supp_text:
            content.append({
                "type": "text",
                "text": f"SUPPLEMENTARY FILES:\n\n{supp_text}",
            })
        content.append({
            "type": "text",
            "text": "Extract all somatic LINE-1 insertion loci from this paper and its supplementary data. Return only valid JSON.",
        })
        messages = [{"role": "user", "content": content}]
    elif supp_text:
        model = MODEL_PDF  # supplements can be large tables — use opus
        messages = [
            {
                "role": "user",
                "content": (
                    f"PAPER ABSTRACT:\n{abstract}\n\n"
                    f"SUPPLEMENTARY FILES:\n\n{supp_text}\n\n"
                    "Extract all somatic LINE-1 insertion loci with genomic coordinates "
                    "from the supplementary data above. Return only valid JSON."
                ),
            }
        ]
        logger.info("PMID %s: using supplements-only mode", pmid)
    else:
        model = MODEL_ABSTRACT
        messages = _build_messages_abstract(abstract)

    # First extraction attempt
    try:
        response = call_claude(client, model, system_blocks, messages, logger)
        total_cost += estimate_cost(response.usage, model)
        raw_text = response.content[0].text if response.content else ""
    except Exception as exc:  # noqa: BLE001
        logger.error("PMID %s: API call failed: %s", pmid, exc)
        return [], "api_error", total_cost

    # Parse JSON
    parsed = extract_json(raw_text)

    if parsed is None:
        logger.warning("PMID %s: JSON parse failed on first attempt; retrying with repair prompt", pmid)
        repair_messages = messages + [
            {"role": "assistant", "content": raw_text},
            {"role": "user", "content": REPAIR_PROMPT},
        ]
        try:
            repair_response = call_claude(client, model, system_blocks, repair_messages, logger)
            total_cost += estimate_cost(repair_response.usage, model)
            raw_text = repair_response.content[0].text if repair_response.content else ""
            parsed = extract_json(raw_text)
        except Exception as exc:  # noqa: BLE001
            logger.error("PMID %s: repair API call failed: %s", pmid, exc)
            parsed = None

    if parsed is None:
        logger.error("PMID %s: JSON parse failed after repair. Marking json_parse_error.", pmid)
        status = "json_parse_error"
        loci: list[dict] = []
    else:
        loci = parsed.get("loci", [])
        reason = parsed.get("reason", "")
        status = "ok"
        logger.info("PMID %s: extracted %d loci | reason: %s", pmid, len(loci), reason[:120])

    # Save raw extraction
    extraction_record = {
        "pmid": pmid,
        "model": model,
        "status": status,
        "cost_usd": round(total_cost, 6),
        "raw_response": raw_text,
        "parsed": parsed,
    }
    extractions_dir.mkdir(parents=True, exist_ok=True)
    extraction_out = extractions_dir / f"{pmid}.json"
    extraction_out.write_text(json.dumps(extraction_record, indent=2, ensure_ascii=False))
    logger.debug("Saved raw extraction: %s", extraction_out)

    return loci, status, total_cost


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract LINE-1 somatic insertion loci from papers using Claude API."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/raw/papers_to_extract.json"),
        help="Input manifest JSON (default: data/raw/papers_to_extract.json).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/raw/loci_raw.json"),
        help="Output flat loci JSON (default: data/raw/loci_raw.json).",
    )
    parser.add_argument(
        "--extractions_dir",
        type=Path,
        default=Path("data/raw/extractions"),
        help="Directory for per-paper raw extraction JSON (default: data/raw/extractions).",
    )
    parser.add_argument(
        "--prompt",
        type=Path,
        default=Path("pipeline/prompts/extract_loci.txt"),
        help="Path to extraction prompt file (default: pipeline/prompts/extract_loci.txt).",
    )
    parser.add_argument(
        "--max_papers",
        type=int,
        default=None,
        help="Cap number of papers to send to Claude (useful for test runs, e.g. --max_papers 5).",
    )
    parser.add_argument(
        "--log_dir",
        type=Path,
        default=Path("logs"),
        help="Directory for log files (default: logs).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    script_name = Path(__file__).stem
    logger = setup_logging(args.log_dir, script_name)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        logger.error("ANTHROPIC_API_KEY environment variable is not set.")
        sys.exit(1)

    # Session header
    logger.info("=== SESSION START: %s ===", script_name)
    logger.info("Date/time       : %s", datetime.now().isoformat())
    logger.info("Python          : %s", sys.version.split()[0])
    logger.info("anthropic SDK   : %s", anthropic.__version__)
    logger.info("Working dir     : %s", Path.cwd())
    logger.info("Log dir         : %s", args.log_dir.resolve())
    logger.info("Input           : %s", args.input.resolve())
    logger.info("Output          : %s", args.output.resolve())
    logger.info("Extractions dir : %s", args.extractions_dir.resolve())
    logger.info("Prompt file     : %s", args.prompt.resolve())
    logger.info("PDF model       : %s", MODEL_PDF)
    logger.info("Abstract model  : %s", MODEL_ABSTRACT)

    # Load system prompt
    system_prompt = load_system_prompt(args.prompt, logger)

    # Load paper manifest
    if not args.input.exists():
        logger.error("Input file not found: %s", args.input.resolve())
        sys.exit(1)

    raw = args.input.read_text(encoding="utf-8")
    papers: list[dict] = json.loads(raw) if raw.strip() else []
    logger.info("Loaded %d papers from %s", len(papers), args.input.resolve())

    if args.max_papers and len(papers) > args.max_papers:
        logger.info("--max_papers %d: capping from %d papers", args.max_papers, len(papers))
        papers = papers[: args.max_papers]

    if not papers:
        logger.info("No papers to process. Writing empty loci file.")
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps([], indent=2))
        logger.info("=== DONE: %s completed successfully ===", script_name)
        return

    client = anthropic.Anthropic(api_key=api_key)

    # Process papers
    all_loci: list[dict] = []
    n_ok = 0
    n_json_errors = 0
    n_api_errors = 0
    total_cost_usd = 0.0

    for idx, paper in enumerate(papers, start=1):
        pmid = paper.get("pmid", "?")
        logger.info(
            "=== Processing paper %d/%d (PMID=%s) ===",
            idx, len(papers), pmid,
        )

        loci, status, cost = process_paper(
            paper, client, system_prompt, args.extractions_dir, logger
        )
        total_cost_usd += cost

        if status == "ok":
            n_ok += 1
            # Annotate each locus with the source PMID
            for locus in loci:
                locus["pmid"] = pmid
            all_loci.extend(loci)
        elif status == "json_parse_error":
            n_json_errors += 1
        else:
            n_api_errors += 1

    # Summary
    logger.info("=== Extraction Summary ===")
    logger.info("Papers processed          : %d", len(papers))
    logger.info("Successful extractions    : %d", n_ok)
    logger.info("JSON parse errors         : %d", n_json_errors)
    logger.info("API errors                : %d", n_api_errors)
    logger.info("Total loci extracted      : %d", len(all_loci))
    logger.info("Estimated total cost (USD): $%.4f", total_cost_usd)

    # Save flat loci file
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(all_loci, indent=2, ensure_ascii=False))
    logger.info(
        "Saved loci_raw.json: %s (%d loci, %.1f KB)",
        args.output.resolve(),
        len(all_loci),
        args.output.stat().st_size / 1024,
    )

    logger.info("=== DONE: %s completed successfully ===", script_name)


if __name__ == "__main__":
    main()
