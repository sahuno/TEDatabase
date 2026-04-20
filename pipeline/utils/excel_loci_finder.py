# Author: Samuel Ahuno
# Date: 2026-04-20
# Purpose: Smart Excel/CSV supplement parser that scores and filters sheets containing genomic loci

from __future__ import annotations

import csv
import logging
from pathlib import Path

# ---------------------------------------------------------------------------
# Scoring constants
# ---------------------------------------------------------------------------

COORD_HEADERS: set[str] = {
    "chr", "chrom", "chromosome", "seqnames", "contig",
    "start", "pos", "position", "begin", "chromstart",
    "end", "stop", "chromend",
}

_CHR_HEADERS: set[str] = {"chr", "chrom", "chromosome", "seqnames", "contig"}
_START_HEADERS: set[str] = {"start", "pos", "position", "begin", "chromstart"}
_END_HEADERS: set[str] = {"end", "stop", "chromend"}

L1_HEADERS: set[str] = {
    "l1", "line1", "line-1", "l1hs", "mei", "insertion",
    "retrotransposon", "transposon", "te", "mobile",
}

_BONUS_HEADERS: set[str] = {"strand", "sample", "tumor", "cancer", "tissue"}

SCORE_THRESHOLD: int = 2


# ---------------------------------------------------------------------------
# score_sheet_headers
# ---------------------------------------------------------------------------

def score_sheet_headers(headers: list[str]) -> int:
    """
    Score a list of column headers for likelihood of containing genomic loci.

    Scoring rules:
    - +2 for any chr/chrom variant match
    - +2 for any start-position variant match
    - +1 for any end-position variant match
    - +2 for any LINE-1 / MEI keyword match
    - +1 for each of: strand, sample, tumor, cancer, tissue

    Parameters
    ----------
    headers : list[str]
        Raw column header strings from a sheet or CSV.

    Returns
    -------
    int
        Total relevance score. Higher scores indicate more likely locus data.

    Examples
    --------
    >>> score_sheet_headers(["chr", "start", "end", "strand", "mei_type"])
    8
    >>> score_sheet_headers(["sample_id", "age", "diagnosis"])
    1
    """
    score = 0
    normalised = [h.lower().strip() for h in headers]

    for h in normalised:
        if h in _CHR_HEADERS:
            score += 2
        if h in _START_HEADERS:
            score += 2
        if h in _END_HEADERS:
            score += 1
        if h in L1_HEADERS:
            score += 2
        if h in _BONUS_HEADERS:
            score += 1

    return score


# ---------------------------------------------------------------------------
# find_loci_sheets
# ---------------------------------------------------------------------------

def find_loci_sheets(path: Path, logger: logging.Logger) -> list[dict]:
    """
    Open an Excel file and score each sheet for genomic locus content.

    Reads the first non-empty row of each sheet as headers, scores it with
    :func:`score_sheet_headers`, and returns a list of scored sheet metadata.

    Parameters
    ----------
    path : Path
        Path to the Excel file (.xlsx, .xls, .xlsm).
    logger : logging.Logger
        Logger instance for audit output.

    Returns
    -------
    list[dict]
        One dict per sheet with keys:
        ``sheet_name`` (str), ``score`` (int), ``headers`` (list[str]),
        ``include`` (bool).
        Returns an empty list if the file cannot be opened.

    Examples
    --------
    >>> # With a real Excel file:
    >>> results = find_loci_sheets(Path("supplement.xlsx"), logger)
    >>> [r["sheet_name"] for r in results if r["include"]]
    ['MEI_calls', 'Somatic_insertions']
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed — cannot parse Excel files.")
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Cannot open Excel file %s: %s", path.name, exc)
        return []

    sheet_results: list[dict] = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            headers: list[str] = []

            # Find the first non-empty row and use it as headers
            for row in ws.iter_rows(values_only=True):
                # A row is non-empty if at least one cell has a value
                non_null = [c for c in row if c is not None and str(c).strip()]
                if non_null:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    break

            score = score_sheet_headers(headers)
            include = score >= SCORE_THRESHOLD
            decision = "INCLUDE" if include else "skip"

            logger.info(
                "Sheet %-30s | score=%2d | %-6s | headers=%s",
                f'"{sheet_name}"',
                score,
                decision,
                headers[:8],
            )

            sheet_results.append({
                "sheet_name": sheet_name,
                "score": score,
                "headers": headers,
                "include": include,
            })

        except Exception as exc:
            logger.warning("Error reading sheet '%s' in %s: %s", sheet_name, path.name, exc)
            sheet_results.append({
                "sheet_name": sheet_name,
                "score": 0,
                "headers": [],
                "include": False,
            })

    try:
        wb.close()
    except Exception:
        pass

    n_included = sum(1 for s in sheet_results if s["include"])
    logger.info(
        "%s: %d/%d sheets will be included (score >= %d)",
        path.name, n_included, len(sheet_results), SCORE_THRESHOLD,
    )
    return sheet_results


# ---------------------------------------------------------------------------
# excel_to_text_smart
# ---------------------------------------------------------------------------

def excel_to_text_smart(
    path: Path,
    logger: logging.Logger,
    max_rows: int = 2000,
) -> str:
    """
    Convert an Excel file to plain text, skipping low-relevance sheets.

    High-scoring sheets (score >= SCORE_THRESHOLD) are fully converted to
    tab-separated text. Low-scoring sheets are represented by a single
    commented-out header line so the caller knows they exist without paying
    token cost for their content.

    Parameters
    ----------
    path : Path
        Path to the Excel file.
    logger : logging.Logger
        Logger instance.
    max_rows : int, optional
        Maximum data rows to include per sheet (default 2000). Header row is
        not counted toward this limit.

    Returns
    -------
    str
        Combined text for all sheets, with section markers.
        Returns an empty string if the file cannot be parsed.

    Examples
    --------
    >>> text = excel_to_text_smart(Path("supplement.xlsx"), logger)
    >>> "=== Sheet:" in text
    True
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed — cannot parse Excel files.")
        return ""

    scored_sheets = find_loci_sheets(path, logger)
    if not scored_sheets:
        return ""

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Cannot re-open Excel file %s for content extraction: %s", path.name, exc)
        return ""

    sections: list[str] = []

    for sheet_meta in scored_sheets:
        sheet_name = sheet_meta["sheet_name"]
        include = sheet_meta["include"]
        headers = sheet_meta["headers"]

        header_line = "\t".join(headers)

        if not include:
            sections.append(
                f"=== Sheet: {sheet_name} (score={sheet_meta['score']}) ===\n"
                f"# SKIPPED SHEET (score too low): {header_line}\n"
            )
            continue

        # Full extraction for included sheets
        try:
            ws = wb[sheet_name]
            rows_out: list[str] = []
            first_data_row = True
            n_rows = 0

            for row in ws.iter_rows(values_only=True):
                # Skip until we find the first non-empty row (the header)
                if first_data_row:
                    non_null = [c for c in row if c is not None and str(c).strip()]
                    if not non_null:
                        continue
                    # This is the header row — already captured in sheet_meta
                    rows_out.append(header_line)
                    first_data_row = False
                    continue

                if n_rows >= max_rows:
                    rows_out.append(f"# ... truncated at {max_rows} rows ...")
                    break

                cells = [str(c) if c is not None else "" for c in row]
                rows_out.append("\t".join(cells))
                n_rows += 1

            logger.info(
                "Sheet '%s': extracted %d data rows", sheet_name, n_rows
            )
            sections.append(
                f"=== Sheet: {sheet_name} (score={sheet_meta['score']}) ===\n"
                + "\n".join(rows_out)
                + "\n"
            )

        except Exception as exc:
            logger.warning(
                "Error extracting content from sheet '%s': %s", sheet_name, exc
            )
            sections.append(
                f"=== Sheet: {sheet_name} ===\n"
                f"# ERROR reading sheet: {exc}\n"
            )

    try:
        wb.close()
    except Exception:
        pass

    return "\n".join(sections)


# ---------------------------------------------------------------------------
# csv_to_text
# ---------------------------------------------------------------------------

def csv_to_text(
    path: Path,
    logger: logging.Logger,
    max_rows: int = 2000,
) -> str:
    """
    Read a CSV or TSV file and return its content as tab-separated text.

    Sniffs the delimiter automatically (comma vs tab). Logs headers and the
    sheet relevance score.

    Parameters
    ----------
    path : Path
        Path to the CSV/TSV file.
    logger : logging.Logger
        Logger instance.
    max_rows : int, optional
        Maximum rows to include (default 2000, excluding header).

    Returns
    -------
    str
        Tab-separated text with a header row, capped at max_rows data rows.
        Returns an empty string on any error.

    Examples
    --------
    >>> text = csv_to_text(Path("mei_calls.tsv"), logger)
    >>> text.startswith("chr\\t")
    True
    """
    try:
        raw = path.read_bytes()
        # Decode with a fallback — supplement files are sometimes latin-1
        try:
            content = raw.decode("utf-8")
        except UnicodeDecodeError:
            content = raw.decode("latin-1")
            logger.info("%s: decoded with latin-1 fallback", path.name)

        # Sniff delimiter from first 4 KB
        sample = content[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if "\t" in sample else ","

        lines = content.splitlines()
        if not lines:
            logger.warning("%s: file appears empty", path.name)
            return ""

        reader = csv.reader(lines, delimiter=delimiter)
        rows_out: list[str] = []
        headers: list[str] = []
        n_rows = 0

        for i, row in enumerate(reader):
            if i == 0:
                headers = [h.strip() for h in row]
                score = score_sheet_headers(headers)
                logger.info(
                    "%s: %d columns | score=%d | delimiter=%r | headers=%s",
                    path.name, len(headers), score, delimiter, headers[:8],
                )
                rows_out.append("\t".join(headers))
                continue

            if n_rows >= max_rows:
                rows_out.append(f"# ... truncated at {max_rows} rows ...")
                break

            rows_out.append("\t".join(row))
            n_rows += 1

        logger.info("%s: %d data rows included", path.name, n_rows)
        return "\n".join(rows_out) + "\n"

    except Exception as exc:
        logger.warning("Cannot read CSV/TSV file %s: %s", path.name, exc)
        return ""


# ---------------------------------------------------------------------------
# supplements_to_text
# ---------------------------------------------------------------------------

def supplements_to_text(
    paths: list[Path],
    logger: logging.Logger,
) -> str:
    """
    Convert a list of supplementary files to a single text block.

    Dispatches each file to the appropriate reader:
    - ``.xlsx`` / ``.xls`` / ``.xlsm``: :func:`excel_to_text_smart`
    - ``.csv`` / ``.tsv`` / ``.txt`` (tabular): :func:`csv_to_text`
    - Other text files: plain read

    Parameters
    ----------
    paths : list[Path]
        List of file paths to process.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    str
        All file contents concatenated, each prefixed with a section header:
        ``--- Supplementary file: {name} ---``

    Examples
    --------
    >>> text = supplements_to_text([Path("table_s1.xlsx"), Path("table_s2.csv")], logger)
    >>> "--- Supplementary file:" in text
    True
    """
    _EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".xlsb"}
    _CSV_SUFFIXES = {".csv", ".tsv", ".txt", ".tab"}

    all_sections: list[str] = []

    for path in paths:
        suffix = path.suffix.lower()
        logger.info("--- Processing supplement: %s (type=%s) ---", path.name, suffix or "unknown")

        try:
            if suffix in _EXCEL_SUFFIXES:
                content = excel_to_text_smart(path, logger)
                file_type = "excel"
            elif suffix in _CSV_SUFFIXES:
                content = csv_to_text(path, logger)
                file_type = "csv/tsv"
            else:
                # Plain text fallback — read up to ~500 KB
                try:
                    raw = path.read_bytes()
                    try:
                        content = raw[:512_000].decode("utf-8")
                    except UnicodeDecodeError:
                        content = raw[:512_000].decode("latin-1")
                    file_type = "text"
                    logger.info(
                        "%s: read as plain text (%d chars)", path.name, len(content)
                    )
                except Exception as exc:
                    logger.warning("Cannot read plain text file %s: %s", path.name, exc)
                    content = ""
                    file_type = "unknown"

            char_count = len(content)
            logger.info(
                "Supplement %s | type=%-8s | chars=%d",
                path.name, file_type, char_count,
            )

            section_header = f"--- Supplementary file: {path.name} ---"
            if content.strip():
                all_sections.append(f"{section_header}\n{content}")
            else:
                all_sections.append(f"{section_header}\n# (no extractable content)\n")

        except Exception as exc:
            logger.warning("Unhandled error processing %s: %s", path.name, exc)
            all_sections.append(
                f"--- Supplementary file: {path.name} ---\n# ERROR: {exc}\n"
            )

    return "\n".join(all_sections)
