# Author: Samuel Ahuno
# Date: 2026-04-20
# Purpose: Seed script to import somatic LINE-1 insertions from the PCAWG consortium

from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Iterator

# Make the project root importable so `pipeline.utils` resolves when run as a script
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from pipeline.utils.coordinates import make_locus_id, parse_chrom

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_NAME = Path(__file__).stem

# Rodriguez-Martin 2020 Nature Genetics — PMID 32024998
SOURCE_PMID = ["32024998"]
SOURCE_DOI = "10.1038/s41588-019-0562-0"
PAPER_TITLE = (
    "Pan-cancer analysis of whole genomes identifies driver rearrangements "
    "promoted by mobile element insertions"
)
PAPER_YEAR = 2020
PAPER_JOURNAL = "Nature Genetics"

# PCAWG used GRCh37 / hg19
GENOME_BUILD = "hg19"

# Keywords that identify LINE-1 rows in MEI call files
_L1_KEYWORDS: frozenset[str] = frozenset({
    "line-1", "line1", "l1", "l1hs", "l1pa", "l1p",
})

# Keywords to EXCLUDE (other MEI classes)
_EXCLUDE_KEYWORDS: frozenset[str] = frozenset({
    "alu", "sva", "herv", "herv-k", "hervk",
})

# Download attempts in priority order
_DOWNLOAD_ATTEMPTS: list[dict] = [
    {
        "label": "ICGC DCC PCAWG driver mutations",
        "url": (
            "https://dcc.icgc.org/api/v1/download?fn=/PCAWG/driver_mutations/"
            "pcawg_mei_driver_mutations.tsv"
        ),
    },
    {
        "label": "ICGC DCC PCAWG MEI TSV (alternate path)",
        "url": (
            "https://dcc.icgc.org/api/v1/download?fn=/PCAWG/driver_mutations/"
            "TableS3_panorama_driver_mutations_in_cancer_genomes.public.tsv"
        ),
    },
    {
        "label": "Rodriguez-Martin 2020 Nature Genetics MOESM3 (correct URL)",
        "url": (
            "https://static-content.springer.com/esm/"
            "art%3A10.1038%2Fs41588-019-0562-0/MediaObjects/"
            "41588_2019_562_MOESM3_ESM.xlsx"
        ),
    },
    {
        "label": "PCAWG GitHub germline-sv repository",
        "url": (
            "https://raw.githubusercontent.com/ICGC-TCGA-PanCancer/"
            "pcawg-germline-sv/master/supplementary_tables/"
            "supplementary_table_mei_calls.tsv"
        ),
    },
]

_MANUAL_INSTRUCTIONS = """
============================================================
MANUAL DOWNLOAD REQUIRED — all automatic sources failed.
============================================================

The PCAWG somatic MEI dataset requires access credentials for some endpoints.
Please follow these steps:

1. ICGC DCC (recommended):
   a. Register at https://dcc.icgc.org/
   b. Navigate to: PCAWG > Files > driver_mutations
   c. Search for files containing "MEI" or "retrotransposon"
   d. Download and provide via --input <path>

2. Synapse (syn12975893):
   a. Install: pip install synapseclient
   b. synapse get syn12975893 -r
   c. Look for MEI-related files and provide via --input <path>

3. Nature Genetics supplementary data:
   DOI: 10.1038/s41588-020-0638-4
   Download MOESM files from the paper's supplementary section.

4. Once you have the file, re-run:
   python pipeline/seed_pcawg.py --input /path/to/mei_calls.tsv

============================================================
"""


# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

def setup_logging(log_dir: Path) -> logging.Logger:
    """
    Configure dual-handler logging: FileHandler (DEBUG) + StreamHandler (INFO).

    Parameters
    ----------
    log_dir : Path
        Directory for the timestamped log file.

    Returns
    -------
    logging.Logger
        Configured logger instance.
    """
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"{SCRIPT_NAME}_{timestamp}.log"

    logger = logging.getLogger(SCRIPT_NAME)
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s")

    fh = logging.FileHandler(log_file)
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(sh)

    logger.info("Log file: %s", log_file.resolve())
    return logger


# ---------------------------------------------------------------------------
# Download helpers
# ---------------------------------------------------------------------------

def _try_download(url: str, timeout: int = 30) -> bytes | None:
    """
    Attempt a single HTTP GET. Return bytes on success, None on any error.

    Parameters
    ----------
    url : str
        URL to fetch.
    timeout : int, optional
        Request timeout in seconds (default 30).

    Returns
    -------
    bytes or None
        Response body, or None if the request failed.
    """
    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (compatible; TEDatabase/1.0; "
                    "research use; contact ekwame001@gmail.com)"
                )
            },
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except (urllib.error.URLError, urllib.error.HTTPError, OSError, TimeoutError):
        return None


def attempt_downloads(logger: logging.Logger) -> tuple[bytes, str] | None:
    """
    Try each download source in priority order.

    Parameters
    ----------
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    tuple[bytes, str] or None
        ``(content_bytes, source_label)`` on first success, else ``None``.
    """
    for attempt in _DOWNLOAD_ATTEMPTS:
        label = attempt["label"]
        url = attempt["url"]
        logger.info("Trying download: %s", label)
        logger.debug("URL: %s", url)

        data = _try_download(url)
        if data is not None and len(data) > 512:
            # Skip HTML error/login pages returned with 200 status
            snippet = data[:256].lower()
            if b"<!doctype" in snippet or b"<html" in snippet:
                logger.info("Download returned HTML page (not data): %s", label)
                continue
            logger.info(
                "Download succeeded: %s (%.2f KB)", label, len(data) / 1024
            )
            return data, label
        else:
            logger.info("Download failed or empty response: %s", label)

    return None


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def _sniff_delimiter(header_line: str) -> str:
    """Return the most likely field delimiter for a header line."""
    tab_count = header_line.count("\t")
    comma_count = header_line.count(",")
    return "\t" if tab_count >= comma_count else ","


def _normalise_header(h: str) -> str:
    """Strip, lowercase, and remove surrounding quotes from a column header."""
    return h.strip().strip('"').strip("'").lower().replace(" ", "_").replace("-", "_")


def _detect_format(headers: list[str]) -> str:
    """
    Identify which of the known PCAWG MEI formats the headers match.

    Parameters
    ----------
    headers : list[str]
        Normalised column header names.

    Returns
    -------
    str
        One of ``"format_a"`` (BED-like), ``"format_b"`` (PCAWG annotated),
        ``"format_c"`` (VCF-derived), or ``"unknown"``.
    """
    header_set = set(headers)

    # Format B — PCAWG annotated
    if "icgc_donor_id" in header_set or "chromosome_start" in header_set:
        return "format_b"

    # Format C — VCF-derived with MEINFO / SVTYPE
    if "svtype" in header_set or "meinfo" in header_set or "info" in header_set:
        return "format_c"

    # Format A — BED-like (chr / start / end columns)
    chr_headers = {"chr", "chrom", "chromosome", "seqnames"}
    start_headers = {"start", "chromstart", "pos", "position", "begin"}
    if header_set & chr_headers and header_set & start_headers:
        return "format_a"

    return "unknown"


# ---------------------------------------------------------------------------
# LINE-1 type detection helpers
# ---------------------------------------------------------------------------

def _is_l1(mei_type: str | None) -> bool:
    """
    Return True if the MEI type string refers to a LINE-1 element.

    Parameters
    ----------
    mei_type : str or None
        Free-text MEI type field (e.g. "LINE-1/L1", "Alu", "L1HS").

    Returns
    -------
    bool

    Examples
    --------
    >>> _is_l1("LINE-1/L1")
    True
    >>> _is_l1("Alu")
    False
    >>> _is_l1(None)
    False
    """
    if not mei_type:
        return False
    lower = mei_type.lower().replace("/", " ").replace("_", " ")
    # Must match at least one L1 keyword and NOT match any exclusion keyword
    has_l1 = any(kw in lower for kw in _L1_KEYWORDS)
    has_exclude = any(kw in lower for kw in _EXCLUDE_KEYWORDS)
    return has_l1 and not has_exclude


def _extract_mei_type_from_info(info: str) -> str | None:
    """
    Parse MEINFO= or SVTYPE= from a VCF INFO field string.

    Parameters
    ----------
    info : str
        VCF INFO column string, e.g. ``"SVTYPE=INS;MEINFO=L1HS,100,200,+"``.

    Returns
    -------
    str or None
        Extracted MEI type, or None if not found.
    """
    for part in info.split(";"):
        if part.startswith("MEINFO="):
            return part.split("=", 1)[1].split(",")[0]
        if part.startswith("SVTYPE="):
            svtype = part.split("=", 1)[1]
            if svtype.upper() not in ("INS", "DEL", "DUP", "INV", "TRA"):
                return svtype  # SVTYPE is the actual element name
    return None


# ---------------------------------------------------------------------------
# Format-specific row parsers
# ---------------------------------------------------------------------------

def _parse_format_a(
    rows: list[dict],
    logger: logging.Logger,
) -> Iterator[dict]:
    """
    Parse BED-like TSV rows (Format A).

    Expected columns: chr/chrom, start, end, mei_type, sample_id, strand.
    Flexible column name matching via normalised header lookup.

    Parameters
    ----------
    rows : list[dict]
        List of dicts from csv.DictReader.
    logger : logging.Logger
        Logger instance.

    Yields
    ------
    dict
        Extracted fields: chrom, start, end, strand, sample_id,
        cancer_type, mei_type.
    """
    if not rows:
        return

    sample_headers = rows[0].keys()
    norm_map = {_normalise_header(h): h for h in sample_headers}

    def _get(row: dict, *candidates: str) -> str | None:
        for c in candidates:
            orig = norm_map.get(c)
            if orig and row.get(orig) not in (None, ""):
                return str(row[orig]).strip()
        return None

    for row in rows:
        chrom_raw = _get(row, "chr", "chrom", "chromosome", "seqnames", "contig")
        start_raw = _get(row, "start", "chromstart", "pos", "position", "begin")
        end_raw = _get(row, "end", "chromend", "stop")
        strand = _get(row, "strand") or "."
        sample_id = _get(row, "sample_id", "sample", "tumor_sample", "donor_id")
        cancer_type = _get(row, "cancer_type", "tumor_type", "histology", "project_code")
        mei_type = _get(row, "mei_type", "type", "element", "family", "subtype", "name")

        if not chrom_raw or not start_raw:
            continue

        try:
            start = int(float(start_raw))
            end = int(float(end_raw)) if end_raw else start + 1
        except (ValueError, TypeError):
            continue

        yield {
            "chrom": chrom_raw,
            "start": start,
            "end": end,
            "strand": strand,
            "sample_id": sample_id,
            "cancer_type": cancer_type,
            "mei_type": mei_type,
        }


def _parse_format_b(
    rows: list[dict],
    logger: logging.Logger,
) -> Iterator[dict]:
    """
    Parse PCAWG annotated format rows (Format B).

    Expected columns: icgc_donor_id, chromosome, chromosome_start,
    chromosome_end, assembly_version, mutation_type, subtype.

    Parameters
    ----------
    rows : list[dict]
        List of dicts from csv.DictReader.
    logger : logging.Logger
        Logger instance.

    Yields
    ------
    dict
        Extracted fields: chrom, start, end, strand, sample_id,
        cancer_type, mei_type.
    """
    if not rows:
        return

    sample_headers = rows[0].keys()
    norm_map = {_normalise_header(h): h for h in sample_headers}

    def _get(row: dict, *candidates: str) -> str | None:
        for c in candidates:
            orig = norm_map.get(c)
            if orig and row.get(orig) not in (None, ""):
                return str(row[orig]).strip()
        return None

    for row in rows:
        chrom_raw = _get(row, "chrom", "chromosome", "chr")
        start_raw = _get(row, "beg", "chromosome_start", "start", "chromstart", "pos")
        end_raw = _get(row, "end", "chromosome_end", "chromend", "stop")
        strand = _get(row, "strand", "strand_plus") or "."
        sample_id = _get(row, "icgc_donor_id", "tumor_wgs_icgc_sample_id", "donor_id", "sample_id")
        cancer_type = _get(row, "histology_abbreviation", "project_code", "cancer_type", "ttype")
        mei_type = _get(row, "family", "mutation_type", "subtype", "type", "subfamily")

        if not chrom_raw or not start_raw:
            continue

        # PCAWG uses 1-based coords — convert to 0-based
        try:
            start = int(float(start_raw)) - 1
            end = int(float(end_raw)) if end_raw else start + 1
        except (ValueError, TypeError):
            continue

        if start < 0:
            start = 0

        yield {
            "chrom": chrom_raw,
            "start": start,
            "end": end,
            "strand": strand,
            "sample_id": sample_id,
            "cancer_type": cancer_type,
            "mei_type": mei_type,
        }


def _parse_format_c(
    rows: list[dict],
    logger: logging.Logger,
) -> Iterator[dict]:
    """
    Parse VCF-derived TSV rows (Format C).

    Handles SVTYPE=INS rows with MEINFO fields. Also handles pre-extracted
    TSV where the INFO field has been expanded into columns.

    Parameters
    ----------
    rows : list[dict]
        List of dicts from csv.DictReader.
    logger : logging.Logger
        Logger instance.

    Yields
    ------
    dict
        Extracted fields: chrom, start, end, strand, sample_id,
        cancer_type, mei_type.
    """
    if not rows:
        return

    sample_headers = rows[0].keys()
    norm_map = {_normalise_header(h): h for h in sample_headers}

    def _get(row: dict, *candidates: str) -> str | None:
        for c in candidates:
            orig = norm_map.get(c)
            if orig and row.get(orig) not in (None, ""):
                return str(row[orig]).strip()
        return None

    for row in rows:
        chrom_raw = _get(row, "chrom", "chr", "#chrom", "chromosome")
        pos_raw = _get(row, "pos", "start", "position")
        end_raw = _get(row, "end", "stop")
        info_str = _get(row, "info") or ""
        svtype = _get(row, "svtype") or ""
        sample_id = _get(row, "sample", "sample_id", "tumor_sample_barcode")
        cancer_type = _get(row, "cancer_type", "project", "tumor_type")

        # Determine MEI type: from MEINFO in INFO, or from dedicated column
        mei_type = _get(row, "meinfo", "me_type", "mei_type")
        if not mei_type and info_str:
            mei_type = _extract_mei_type_from_info(info_str)
        if not mei_type and svtype and svtype.upper() not in ("INS",):
            mei_type = svtype

        # Extract strand from MEINFO if present: MEINFO=name,start,end,strand
        strand = _get(row, "strand") or "."
        if strand == "." and "MEINFO=" in info_str:
            meinfo_val = ""
            for part in info_str.split(";"):
                if part.startswith("MEINFO="):
                    meinfo_val = part.split("=", 1)[1]
            parts = meinfo_val.split(",")
            if len(parts) >= 4 and parts[3] in ("+", "-"):
                strand = parts[3]

        if not chrom_raw or not pos_raw:
            continue

        try:
            start = int(float(pos_raw)) - 1  # VCF is 1-based
            if start < 0:
                start = 0
            end = int(float(end_raw)) if end_raw else start + 1
        except (ValueError, TypeError):
            continue

        yield {
            "chrom": chrom_raw,
            "start": start,
            "end": end,
            "strand": strand,
            "sample_id": sample_id,
            "cancer_type": cancer_type,
            "mei_type": mei_type,
        }


# ---------------------------------------------------------------------------
# Cancer type normalisation
# ---------------------------------------------------------------------------

# PCAWG project codes -> human-readable cancer types
_PCAWG_PROJECT_MAP: dict[str, str] = {
    "BLCA": "Bladder-TCC",
    "BRCA": "Breast-AdenoCA",
    "CESC": "Cervix-SCC",
    "CLL": "Lymph-CLL",
    "COAD": "ColoRect-AdenoCA",
    "ESCA": "Eso-AdenoCa",
    "GBM": "Brain-GBM",
    "HNSC": "Head-SCC",
    "KIRC": "Kidney-RCC",
    "KIRP": "Kidney-RCC",
    "LAML": "Blood-AML",
    "LIHC": "Liver-HCC",
    "LUAD": "Lung-AdenoCA",
    "LUSC": "Lung-SCC",
    "OV": "Ovary-AdenoCA",
    "PAAD": "Panc-AdenoCA",
    "PRAD": "Prost-AdenoCA",
    "READ": "ColoRect-AdenoCA",
    "SKCM": "Skin-Melanoma",
    "STAD": "Stomach-AdenoCA",
    "THCA": "Thy-AdenoCA",
    "UCEC": "Uterus-AdenoCA",
}


def normalize_cancer_type(raw: str | None) -> str | None:
    """
    Normalise a cancer type string to a consistent label.

    Maps TCGA/PCAWG project codes to human-readable PCAWG histology names.
    Returns the original value (stripped) if no mapping exists.

    Parameters
    ----------
    raw : str or None
        Raw cancer type / project code string.

    Returns
    -------
    str or None
        Normalised cancer type string, or None if input is None/empty.

    Examples
    --------
    >>> normalize_cancer_type("BRCA")
    'Breast-AdenoCA'
    >>> normalize_cancer_type("Lung-AdenoCA")
    'Lung-AdenoCA'
    >>> normalize_cancer_type(None) is None
    True
    """
    if not raw:
        return None
    stripped = raw.strip()
    upper = stripped.upper()
    # Check direct TCGA code match
    if upper in _PCAWG_PROJECT_MAP:
        return _PCAWG_PROJECT_MAP[upper]
    # Check if it starts with a known project code (e.g. "BRCA-US")
    for code, label in _PCAWG_PROJECT_MAP.items():
        if upper.startswith(code + "-") or upper.startswith(code + "_"):
            return label
    return stripped if stripped else None


# ---------------------------------------------------------------------------
# TSV/CSV parsing from raw bytes
# ---------------------------------------------------------------------------

def parse_mei_bytes(
    data: bytes,
    source_label: str,
    logger: logging.Logger,
) -> list[dict]:
    """
    Decode raw bytes as a TSV/CSV MEI file and parse all rows.

    Tries UTF-8 first, falls back to latin-1. Auto-detects delimiter.
    Detects the file format and dispatches to the appropriate parser.
    Filters to keep only LINE-1 rows.

    Parameters
    ----------
    data : bytes
        Raw file content.
    source_label : str
        Human-readable label for the source (used in log messages).
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    list[dict]
        List of parsed rows, each with keys: chrom, start, end, strand,
        sample_id, cancer_type, mei_type.
    """
    # Decode
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
        logger.info("%s: decoded with latin-1 fallback", source_label)

    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        logger.warning("%s: file decoded but contains no non-empty lines", source_label)
        return []

    # Skip comment/browser/track lines to find the real header
    header_idx = 0
    for i, line in enumerate(lines):
        if not (line.startswith("#") or line.startswith("browser") or line.startswith("track")):
            header_idx = i
            break

    # Sniff delimiter from header
    header_line = lines[header_idx]
    delimiter = "\t" if header_line.count("\t") >= header_line.count(",") else ","

    reader = csv.DictReader(
        lines[header_idx:],
        delimiter=delimiter,
        quoting=csv.QUOTE_MINIMAL,
    )

    try:
        rows = list(reader)
    except Exception as exc:
        logger.error("%s: CSV parsing error: %s", source_label, exc)
        return []

    if not rows:
        logger.warning("%s: CSV reader produced 0 rows", source_label)
        return []

    headers_raw = list(rows[0].keys())
    headers_norm = [_normalise_header(h) for h in headers_raw]
    fmt = _detect_format(headers_norm)

    logger.info(
        "%s: %d raw rows | delimiter=%r | format=%s | headers=%s",
        source_label, len(rows), delimiter, fmt, headers_raw[:10],
    )

    # Dispatch to format-specific parser
    if fmt == "format_a":
        parsed = list(_parse_format_a(rows, logger))
    elif fmt == "format_b":
        parsed = list(_parse_format_b(rows, logger))
    elif fmt == "format_c":
        parsed = list(_parse_format_c(rows, logger))
    else:
        # Unknown format: try format_a as a best-effort fallback
        logger.warning(
            "%s: unknown format, attempting Format A (BED-like) as fallback", source_label
        )
        parsed = list(_parse_format_a(rows, logger))

    logger.info("%s: %d rows after format parsing", source_label, len(parsed))

    # Filter to LINE-1 only
    l1_rows = [r for r in parsed if _is_l1(r.get("mei_type"))]
    non_l1 = len(parsed) - len(l1_rows)
    logger.info(
        "%s: LINE-1 filter: %d retained, %d removed (Alu/SVA/HERV/other)",
        source_label, len(l1_rows), non_l1,
    )

    return l1_rows


def parse_mei_excel(
    data: bytes,
    source_label: str,
    logger: logging.Logger,
) -> list[dict]:
    """
    Decode raw bytes as an Excel file and extract MEI rows from all sheets.

    Uses openpyxl to read each sheet. Falls back to empty list if openpyxl
    is not installed or the data is not valid Excel.

    Parameters
    ----------
    data : bytes
        Raw Excel file content.
    source_label : str
        Human-readable label for the source (used in log messages).
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    list[dict]
        List of parsed LINE-1 rows across all sheets.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — cannot parse Excel supplement.")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("%s: cannot open as Excel workbook: %s", source_label, exc)
        return []

    all_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows: list[list] = []

        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                sheet_rows.append([str(c) if c is not None else "" for c in row])

        if len(sheet_rows) < 2:
            continue

        # Find the first row with many non-null cells — skip title/description rows
        # (This table has 4 rows of title/description before the actual column headers)
        header_row_idx = 0
        max_non_null = 0
        for idx, row in enumerate(sheet_rows[:10]):
            non_null = sum(1 for c in row if c.strip())
            if non_null > max_non_null:
                max_non_null = non_null
                header_row_idx = idx

        headers = sheet_rows[header_row_idx]
        tsv_lines = ["\t".join(headers)]
        for row in sheet_rows[header_row_idx + 1:]:
            tsv_lines.append("\t".join(row))

        sheet_bytes = "\n".join(tsv_lines).encode("utf-8")
        sheet_parsed = parse_mei_bytes(
            sheet_bytes, f"{source_label}[sheet={sheet_name}]", logger
        )
        all_rows.extend(sheet_parsed)

    try:
        wb.close()
    except Exception:
        pass

    return all_rows


# ---------------------------------------------------------------------------
# Locus construction
# ---------------------------------------------------------------------------

def build_locus(row: dict, today: str) -> dict | None:
    """
    Build a master-format locus dict from a parsed MEI row.

    Parameters
    ----------
    row : dict
        Parsed row with keys: chrom, start, end, strand, sample_id,
        cancer_type, mei_type.
    today : str
        ISO date string (YYYY-MM-DD) for date_added / date_updated.

    Returns
    -------
    dict or None
        Locus dict in master-format, or ``None`` if the chromosome cannot
        be normalised or coordinates are invalid.

    Examples
    --------
    >>> locus = build_locus(
    ...     {"chrom": "1", "start": 12345, "end": 12346,
    ...      "strand": "+", "cancer_type": "BRCA", "mei_type": "L1HS"},
    ...     "2026-04-20",
    ... )
    >>> locus["chrom"]
    'chr1'
    """
    chrom = parse_chrom(str(row.get("chrom") or ""))
    if chrom is None:
        return None

    start = row.get("start")
    end = row.get("end")

    # Validate numeric coordinates
    try:
        start = int(start)
        end = int(end)
    except (TypeError, ValueError):
        return None

    if start < 0 or end <= start:
        end = start + 1  # minimal single-bp interval

    strand_raw = str(row.get("strand") or ".").strip()
    strand = strand_raw if strand_raw in ("+", "-") else "unknown"

    cancer_type_raw = row.get("cancer_type")
    normalised_cancer = normalize_cancer_type(cancer_type_raw)

    locus_id = make_locus_id(chrom, start, end, strand)

    return {
        "locus_id": locus_id,
        "chrom": chrom,
        "start": start,
        "end": end,
        "strand": strand,
        "genome_build": GENOME_BUILD,
        "l1_family": "L1HS",
        "l1_subtype": None,
        "insertion_type": "somatic",
        "cancer_type": [normalised_cancer] if normalised_cancer else [],
        "tissue_type": [],
        "source_type": "literature",
        "source_pmid": list(SOURCE_PMID),
        "source_doi": SOURCE_DOI,
        "paper_title": PAPER_TITLE,
        "paper_year": PAPER_YEAR,
        "paper_journal": PAPER_JOURNAL,
        "coordinate_source": "pcawg_catalog",
        "coordinate_confidence": "high",
        "validation_level": "computational",
        "detection_method": ["WGS"],
        "n_samples_detected": 1,
        "date_added": today,
        "date_updated": today,
        "version": 1,
    }


# ---------------------------------------------------------------------------
# Deduplication with cancer_type merging
# ---------------------------------------------------------------------------

def deduplicate_loci(loci: list[dict], logger: logging.Logger) -> list[dict]:
    """
    Deduplicate loci by locus_id, merging cancer_type lists and summing
    n_samples_detected for shared insertions.

    Parameters
    ----------
    loci : list[dict]
        All locus dicts, possibly containing duplicate locus_ids from
        multiple samples sharing the same insertion.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    list[dict]
        Deduplicated list. Each unique locus_id appears exactly once;
        cancer_type is the union of all contributing records.
    """
    index: dict[str, dict] = {}

    for locus in loci:
        lid = locus["locus_id"]
        if lid not in index:
            index[lid] = locus.copy()
            index[lid]["cancer_type"] = list(locus.get("cancer_type") or [])
        else:
            existing = index[lid]
            # Union-merge cancer types
            merged_ct = sorted(
                set(existing.get("cancer_type") or [])
                | set(locus.get("cancer_type") or [])
            )
            existing["cancer_type"] = merged_ct
            # Sum sample counts
            existing["n_samples_detected"] = (
                (existing.get("n_samples_detected") or 1)
                + (locus.get("n_samples_detected") or 1)
            )

    deduped = list(index.values())
    logger.info(
        "Deduplication: %d total rows -> %d unique loci (%d merged across samples)",
        len(loci), len(deduped), len(loci) - len(deduped),
    )
    return deduped


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def atomic_write_json(path: Path, data: object, logger: logging.Logger) -> None:
    """
    Write a JSON file atomically via a .tmp intermediate.

    Parameters
    ----------
    path : Path
        Target file path.
    data : object
        JSON-serialisable object.
    logger : logging.Logger
        Logger instance.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp.rename(path)
    size_kb = path.stat().st_size / 1024
    logger.info("Wrote (atomic): %s (%.1f KB)", path.resolve(), size_kb)


def write_summary_tsv(
    loci: list[dict],
    path: Path,
    logger: logging.Logger,
) -> None:
    """
    Write a human-readable TSV summary of the seeded loci.

    Parameters
    ----------
    loci : list[dict]
        Deduplicated locus list.
    path : Path
        Output .tsv file path.
    logger : logging.Logger
        Logger instance.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = [
        "locus_id", "chrom", "start", "end", "strand",
        "genome_build", "l1_family", "cancer_type",
        "n_samples_detected", "coordinate_confidence",
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        fh.write("#" + "\t".join(columns) + "\n")
        for locus in loci:
            row = []
            for col in columns:
                val = locus.get(col)
                if isinstance(val, list):
                    row.append("|".join(str(v) for v in val) if val else "")
                elif val is None:
                    row.append("")
                else:
                    row.append(str(val))
            fh.write("\t".join(row) + "\n")

    logger.info("Saved summary TSV: %s (%d loci)", path.resolve(), len(loci))


def load_master(path: Path, logger: logging.Logger) -> list[dict]:
    """
    Load the master loci.json file. Returns an empty list if absent or empty.

    Parameters
    ----------
    path : Path
        Path to master JSON file.
    logger : logging.Logger
        Logger instance.

    Returns
    -------
    list[dict]
        Existing master loci, or empty list.
    """
    if not path.exists():
        logger.info("Master not found — starting fresh: %s", path)
        return []
    raw = path.read_text(encoding="utf-8").strip()
    if not raw or raw in ("{}", "[]"):
        logger.info("Master is empty — starting fresh.")
        return []
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            logger.info("Loaded master: %d existing loci from %s", len(data), path)
            return data
        logger.warning("Master is not a list (%s) — starting fresh.", type(data).__name__)
        return []
    except json.JSONDecodeError as exc:
        logger.error("Cannot parse master JSON (%s): %s — starting fresh.", path, exc)
        return []


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Seed somatic LINE-1 insertions from the PCAWG consortium "
            "(Rodriguez-Martin et al., Nature Genetics 2020) into loci_raw.json."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=(
            "Path to a locally-supplied MEI file (TSV/CSV/Excel/BED). "
            "If provided, all download attempts are skipped."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/raw/loci_raw.json"),
        help="Destination JSON file for raw loci (default: data/raw/loci_raw.json).",
    )
    parser.add_argument(
        "--seed_dir",
        type=Path,
        default=Path("data/seed"),
        help="Directory for cached download files (default: data/seed).",
    )
    parser.add_argument(
        "--master",
        type=Path,
        default=Path("data/processed/loci.json"),
        help="Master loci.json for dedup check (default: data/processed/loci.json).",
    )
    parser.add_argument(
        "--log_dir",
        type=Path,
        default=Path("logs"),
        help="Directory for log files (default: logs).",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    start_time = time.monotonic()
    args = parse_args()
    logger = setup_logging(args.log_dir)

    # Session header
    logger.info("=== SESSION START: %s ===", SCRIPT_NAME)
    logger.info("Date/time   : %s", datetime.now().isoformat())
    logger.info("Python      : %s", sys.version.split()[0])
    logger.info("Working dir : %s", Path.cwd())
    logger.info("Input file  : %s", args.input or "(auto-download)")
    logger.info("Output      : %s", args.output.resolve())
    logger.info("Seed dir    : %s", args.seed_dir.resolve())
    logger.info("Master      : %s", args.master.resolve())
    logger.info("Log dir     : %s", args.log_dir.resolve())
    logger.info("Genome build: %s (PCAWG standard — liftover to hg38 in stage 4)", GENOME_BUILD)

    today = str(date.today())
    raw_data: bytes | None = None
    source_label: str = "local"
    is_excel: bool = False

    # ---------------------------------------------------------------- input source
    if args.input is not None:
        # User supplied a local file — skip downloads
        input_path = args.input
        if not input_path.exists():
            logger.error("--input file not found: %s", input_path.resolve())
            sys.exit(1)
        raw_data = input_path.read_bytes()
        source_label = input_path.name
        is_excel = input_path.suffix.lower() in (".xlsx", ".xls", ".xlsm", ".xlsb")
        logger.info(
            "Using local file: %s (%.2f KB)", input_path.resolve(), len(raw_data) / 1024
        )
    else:
        # Attempt downloads in priority order
        logger.info("=== Attempting automatic downloads ===")
        result = attempt_downloads(logger)

        if result is None:
            # All downloads failed — print manual instructions and exit clean
            logger.warning("All download attempts failed.")
            print(_MANUAL_INSTRUCTIONS)
            logger.info("Exiting without error — manual download required.")
            logger.info("=== DONE: %s completed successfully ===", SCRIPT_NAME)
            sys.exit(0)

        raw_data, source_label = result
        # Detect if the downloaded content is Excel by magic bytes
        is_excel = raw_data[:4] in (
            b"PK\x03\x04",  # ZIP / xlsx
            b"\xd0\xcf\x11\xe0",  # Compound Doc / old xls
        )
        # Cache the downloaded file for reproducibility
        args.seed_dir.mkdir(parents=True, exist_ok=True)
        ext = ".xlsx" if is_excel else ".tsv"
        cached_path = args.seed_dir / f"pcawg_mei_download{ext}"
        cached_path.write_bytes(raw_data)
        logger.info("Cached download: %s (%.2f KB)", cached_path.resolve(), len(raw_data) / 1024)

    # ---------------------------------------------------------------- parse
    logger.info("=== Parsing MEI file: %s ===", source_label)

    if is_excel:
        parsed_rows = parse_mei_excel(raw_data, source_label, logger)
    else:
        parsed_rows = parse_mei_bytes(raw_data, source_label, logger)

    logger.info("Total LINE-1 rows parsed: %d", len(parsed_rows))

    if not parsed_rows:
        logger.warning(
            "No LINE-1 rows extracted from %s. "
            "The file may use an unrecognised format or contain no L1 entries. "
            "Exiting without modifying output.",
            source_label,
        )
        logger.info("=== DONE: %s completed successfully ===", SCRIPT_NAME)
        sys.exit(0)

    # ---------------------------------------------------------------- build loci
    logger.info("=== Building locus records ===")
    all_loci: list[dict] = []
    n_invalid = 0

    for row in parsed_rows:
        locus = build_locus(row, today)
        if locus is None:
            n_invalid += 1
        else:
            all_loci.append(locus)

    logger.info("Loci built: %d valid, %d invalid chrom/coord (skipped)", len(all_loci), n_invalid)

    # ---------------------------------------------------------------- dedup
    logger.info("=== Deduplicating by locus_id ===")
    deduped_loci = deduplicate_loci(all_loci, logger)

    # Summary stats
    cancer_types_seen: set[str] = set()
    for locus in deduped_loci:
        cancer_types_seen.update(locus.get("cancer_type") or [])
    logger.info("Unique cancer types represented: %d", len(cancer_types_seen))
    logger.info("Cancer types: %s", sorted(cancer_types_seen))

    max_samples = max((l.get("n_samples_detected") or 0) for l in deduped_loci) if deduped_loci else 0
    logger.info("Most-recurrent insertion: %d samples", max_samples)

    # ---------------------------------------------------------------- dedup against master
    logger.info("=== Checking against existing master (%s) ===", args.master)
    master_loci = load_master(args.master, logger)
    master_ids: set[str] = {
        l["locus_id"] for l in master_loci if l.get("locus_id")
    }

    new_loci = [l for l in deduped_loci if l["locus_id"] not in master_ids]
    already_in_master = len(deduped_loci) - len(new_loci)
    logger.info(
        "PCAWG loci already in master: %d | New to add: %d",
        already_in_master, len(new_loci),
    )

    # ---------------------------------------------------------------- write outputs
    logger.info("=== Writing outputs ===")

    # Primary output: data/raw/loci_raw.json (overwrite — feeds into stage 4)
    atomic_write_json(args.output, deduped_loci, logger)

    # Summary TSV
    summary_tsv = args.seed_dir / "pcawg_mei_summary.tsv"
    write_summary_tsv(deduped_loci, summary_tsv, logger)

    # Report
    elapsed = time.monotonic() - start_time
    logger.info("=== Summary ===")
    logger.info("Source          : %s", source_label)
    logger.info("Genome build    : %s", GENOME_BUILD)
    logger.info("Total loci      : %d", len(deduped_loci))
    logger.info("New vs master   : %d", len(new_loci))
    logger.info("Primary output  : %s", args.output.resolve())
    logger.info("Summary TSV     : %s", summary_tsv.resolve())
    logger.info("Completed in    : %.1f s", elapsed)

    logger.info("=== DONE: %s completed successfully ===", SCRIPT_NAME)


if __name__ == "__main__":
    main()
